from openpyxl import load_workbook

def scan_xlsx(file_path):

    findings = []

    try:

        wb = load_workbook(
            file_path,
            data_only=False
        )

        findings.append({
            "type": "Sheets",
            "value": len(wb.sheetnames)
        })

        for sheet in wb.sheetnames:

            findings.append({
                "type": "Sheet",
                "value": sheet
            })

    except Exception as e:

        findings.append({
            "type": "Error",
            "value": str(e)
        })

    return findings